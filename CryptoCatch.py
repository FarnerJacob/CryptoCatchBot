from openpyxl import load_workbook
from TweetCollector import collectTweets, getTweetHistory
from Analyze import analyzeOldTweets, pushNotification
from Query import getTwitter, createKeyWord, addKeyWordRow, createTwitter, addTwitterRow, createNotification, createCryptocurrency, addCryptocurrencyRow, createExchange, createWordPair, addWordPairRow
import psycopg2
import time
import threading
from pynput import keyboard

# This adds the ability to stop the program once it starts running by hitting any key
c = threading.Condition()
def on_press(key):
    try:
        global close
        c.acquire()
        close = True
        print("Shutting down...")
        c.notify_all()
        c.release()
        return False
    except AttributeError:
        print('special key {0} pressed'.format(
            key))

def on_release(key):
    print('Shutting down...'.format(
        key))
    if key == keyboard.Key.esc:
        # Stop listener
        return False

#TODO: implement stored lists of twitter names and cryptocurrency tickers
#TODO: implement a system to learn about new cryptocurrencies from cryptocompare
#TODO: implement a web scraper for exchanges checking both available currencies and news letters
#IDEA: as the list of sources to check grows it might be necessary to implement threads for each source and set priorities along with timer to recheck them
#TODO: implement a set time to analyze tweets tagged with key words
#TODO: Change the deletion of beginning and ending characters on tweets to a variable system
#TODO: Create a different notification for tweeted links

# Close variable is the main thread to detect whether all other operations are closed
close = False
# This main method runs a loop to collect new tweets to analyze
if __name__ == '__main__':
    myConnection = psycopg2.connect("dbname='REMOVED' user='REMOVED' password='REMOVED'")
    listener = keyboard.Listener(on_press=on_press)
    listener.start()


    #enter one time code here (this is used for initial setup; it sets up all of the necessary tables for the program to run)

    #end one time code here

    while True:
        cur = getTwitter()
        for name in cur.fetchall() :
            if(not close):
                collectTweets(name)
                time.sleep(1)
        if(close):
            break


'''Code to setup database from the excel files
createExchange()
createCryptocurrency()
createTwitter()
wb = load_workbook('CryptocurrencyList.xlsx')
ws = wb['CryptocurrencyTwitter']
rowLimit = ws.max_row
count = 2
for cells in ws['A2:A'+str(rowLimit)]:
    for cell in cells:
        addTwitterRow(cell.value)
        ticker = ws['B'+str(count)]
        addCryptocurrencyRow(ticker.value, cell.value)
        count+=1
createKeyWord()
keyWords = [' lists ','listed on',' lightning ',' atomic ','atomic swap',' finished',' completed',' now ',' is ']
for word in keyWords:
    addKeyWordRow(word)
createWordPair()
addWordPairRow('listed on',' is ')
addWordPairRow('listed on',' now ')
addWordPairRow(' lightning ',' completed')
addWordPairRow(' lightning ',' finished')
addWordPairRow(' atomic ',' finished')
addWordPairRow(' atomic ',' finished')
addWordPairRow('atomic swap',' finished')
addWordPairRow('atomic swap',' completed')
addWordPairRow(' atomic ', ' now ')
createNotification()
getTweetHistory()
'''
